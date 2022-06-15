# importing openpyxl module
import openpyxl
from pip._vendor.distlib.compat import raw_input



def read_user_input():
    """get input from user"""
    return str(raw_input("Enter word: "))



def open_excel_file(filename):
    """open excel file and return workbook"""
    return openpyxl.load_workbook(filename)


def find_etymology(entry, filename):
    """find etymology"""
    wb_obj = open_excel_file(filename)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    retrieved_objects = []
    for i in range(1, m_row + 1):
        check_against_prefix = sheet_obj.cell(row=i, column=1)
        # split cell contents into separate words)
        split_results = str(check_against_prefix.value).split()

        # if any(str(check_against_prefix.value)) in entry or str(check_against_prefix.value) == entry:
        if any(RESULT in entry for RESULT in split_results):
            meaning = sheet_obj.cell(row=i, column=2)
            origin = sheet_obj.cell(row=i, column=3)
            examples = sheet_obj.cell(row=i, column=4)
            retrieved = {"Root:": check_against_prefix.value, "Meaning:": meaning.value, "Origin:": origin.value,
                         "Examples:": examples.value}
            retrieved_objects.append(retrieved)
    return retrieved_objects


def print_etymology(list_of_retrieves):
    for r in list_of_retrieves:
        for key, value in r.items():
            print(f'{key:20}{value}')
        print("\n")


def run_program():
    """run program"""
    entry = read_user_input()
    r = find_etymology(entry, "suff-prefix_database.xlsx")
    print_etymology(r)


run_program()

